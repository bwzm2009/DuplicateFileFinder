import sys
import os
import hashlib
import subprocess
import send2trash
import zipfile
from PyQt5.QtWidgets import QMessageBox, QInputDialog, QLineEdit, QMenu, QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QPushButton, QLabel, QTreeWidget, QTreeWidgetItem, QHeaderView, QListWidget, QFileDialog, QLineEdit, QHBoxLayout, QRadioButton, QButtonGroup, QProgressBar, QLabel, QAbstractItemView, QListWidgetItem, QTableWidgetItem, QTableWidget, QCheckBox, QTextEdit, QFileSystemModel, QTreeView, QSplitter, QSpacerItem, QSizePolicy
from PyQt5.QtGui import QCursor, QFont
from PyQt5.QtCore import Qt, QThread, QDir, QFile, QTimer, pyqtSignal

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class MainWindow(QMainWindow):
    VERSION = "0.4"
    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.resize(1200, 700)  # Change the numbers as per your requirement

        self.setWindowTitle(f"Duplicate File Finder v{self.VERSION}")

        tabwidget = QTabWidget()
        self.search_directories_tab = SearchDirectoriesTab()  # Store this instance in a variable
        self.delete_options_tab = DeleteOptionsTab()  # Store this instance in a variable
        self.search_criteria_tab = SearchCriteriaTab()  # Store this instance in a variable
        self.duplicates_tab = DuplicatesTab(self.search_directories_tab, self.delete_options_tab, self.search_criteria_tab)  # Pass it to DuplicatesTab
        self.help_tab = HelpTab(self.VERSION)

        tabwidget.addTab(self.duplicates_tab, "Duplicates")
        tabwidget.addTab(self.search_directories_tab, "Search Directories")
        tabwidget.addTab(self.search_criteria_tab, "Search Criteria")
        tabwidget.addTab(self.delete_options_tab, "Delete Options")       
        tabwidget.addTab(self.help_tab, "Help")        
        self.setCentralWidget(tabwidget)

class DuplicatesFinderThread(QThread):
    progress_signal = pyqtSignal(int)  # Signal to emit progress (if needed)
    finished_signal = pyqtSignal(list)  # Signal to notify when the task is done with results

    def __init__(self, directories, search_criteria, get_hash_method):
        super().__init__()
        self.directories = directories
        self.search_criteria = search_criteria
        self.get_hash = get_hash_method

    def run(self):
        file_dict = {}
        search_inside_zip = self.search_criteria.get('search_inside_zip', False)
        
        min_size = self.search_criteria['min_file_size']
        max_size = self.search_criteria['max_file_size']
        allowed_extensions = self.search_criteria.get('file_extensions', set())
        skip_extensions = set(self.search_criteria.get('skip_extensions', []))

        if not self.directories:
            self.finished_signal.emit([])
            return

        for directory, _ in self.directories:
            for foldername, subfolders, filenames in os.walk(directory):
                for filename in filenames:
                    file_path = os.path.join(foldername, filename)
                    file_size = os.path.getsize(file_path) // 1024  # Get file size in KB
                    file_extension = os.path.splitext(filename)[1][1:].lower()  # Convert to lowercase for consistent checking

                    # Checks before processing
                    if not (min_size <= file_size <= max_size):
                        continue
                    if allowed_extensions and file_extension not in allowed_extensions:
                        continue
                    if file_extension in skip_extensions:
                        continue

                    # Regular file handling
                    file_dict[file_path] = self.get_hash(file_path)

                    # ZIP file handling
                    if search_inside_zip and file_extension == 'zip':
                        with zipfile.ZipFile(file_path, 'r') as zip_ref:
                            for zip_info in zip_ref.infolist():
                                with zip_ref.open(zip_info.filename, 'r') as file_in_zip:
                                    file_content = file_in_zip.read()
                                    hash_of_file = hashlib.sha256(file_content).hexdigest()
                                    file_dict[f"{file_path}_inside_zip/{zip_info.filename}"] = hash_of_file

        hashes = {}
        for file, hash in file_dict.items():
            hashes.setdefault(hash, []).append(file)

        duplicates = [files for files in hashes.values() if len(files) > 1]
        self.finished_signal.emit(duplicates)



class NumericTreeWidgetItem(QTreeWidgetItem):
    def __lt__(self, other):
        column = self.treeWidget().sortColumn()

        # Check if the column we're sorting is the file size column (column 3 in your case)
        if column == 3:
            return float(self.text(column)) < float(other.text(column))
        
        return super(NumericTreeWidgetItem, self).__lt__(other)

class DuplicatesTab(QWidget):
    def __init__(self, search_directories_tab, delete_options_tab, search_criteria_tab, *args, **kwargs):  # Add search_criteria_tab argument
        super(DuplicatesTab, self).__init__(*args, **kwargs)
        self.search_directories_tab = search_directories_tab  # Save reference to the SearchDirectoriesTab
        self.delete_options_tab = delete_options_tab  # Save reference to the DeleteOptionsTab
        self.search_criteria_tab = search_criteria_tab  # Save reference to the SearchCriteriaTab
        self.progress = QProgressBar(self)
        self.progress_label = QLabel("Searching...", self)
        self.progress_label.setStyleSheet("color: red; background-color: yellow;")
        self.progress_label.setMinimumSize(100, 20)
        self.progress_label.hide()
        self.hashes = {}
        self.selected_rows = set()

        layout = QVBoxLayout()

        # Set up a timer to toggle the visibility of the progress_label
        self.flashing_timer = QTimer(self)
        self.flashing_timer.timeout.connect(self.toggle_searching_visibility)
        self.flashing_timer.setInterval(500)  # Toggle visibility every 500 milliseconds

        # Create the horizontal layout for the "Export to Excel" button
        export_layout = QHBoxLayout()
        # Add the "Searching..." label to the layout
        export_layout.addWidget(self.progress_label)
        # Add a stretch to push the button to the right
        export_layout.addStretch()
        # Create and setup the "Export to Excel" button
        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        export_layout.addWidget(self.export_button)

        # Add the horizontal layout to the main layout
        layout.addLayout(export_layout)
        
        # Create the treeview
        self.tree = QTreeWidget()
        self.tree.setSortingEnabled(True)  # Allow sorting in QTreeWidget
        self.tree.setHeaderLabels(["Filename", "Duplicate ID", "File Path", "File Size (KB)", "Percent Similar"])
        layout.addWidget(self.tree)
        self.tree.setColumnWidth(0, 300)  # Change the number as per your requirement
        self.tree.setColumnWidth(2, 300)

        # Create the buttons
        self.find_button = QPushButton("Find Duplicate Files")
        self.find_button.clicked.connect(self.find_duplicates)
        layout.addWidget(self.find_button)

        self.delete_button = QPushButton("Delete Selected Files")
        self.delete_button.clicked.connect(self.delete_selected)
        layout.addWidget(self.delete_button)
        
        self.deselect_all_button = QPushButton("Deselect All", self)
        self.deselect_all_button.clicked.connect(self.deselect_all)
        layout.addWidget(self.deselect_all_button)

        self.clear_button = QPushButton("Clear TreeView")
        self.clear_button.clicked.connect(self.clear_treeview)
        layout.addWidget(self.clear_button)
        
        self.total_files_label = QLabel("Total files: 0")
        layout.addWidget(self.total_files_label)
        self.duplicate_files_label = QLabel("Duplicate files: 0")
        layout.addWidget(self.duplicate_files_label)
        self.space_to_free_label = QLabel("Space to be freed: 0.00 MB")
        layout.addWidget(self.space_to_free_label)

        self.setLayout(layout)

#        self.tree.selectionChanged.connect(self.save_selection)  # Moved this line inside __init__ method
#        self.tree.selectionChanged.connect(self.clear_selection)  # Moved this line inside __init__ method

        self.tree.customContextMenuRequested.connect(self.handle_context_menu)
        self.tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
              
    def find_duplicates(self):
        # Fetch the directories list here
        directories = self.search_directories_tab.get_directories()

        if not directories:  # Optionally, you can check if directories are provided
            QMessageBox.warning(self, "Warning", "No directories selected.")
            return

        search_criteria = {
            'min_file_size': self.search_criteria_tab.get_min_file_size(),
            'max_file_size': self.search_criteria_tab.get_max_file_size(),
            'file_extensions': self.search_criteria_tab.get_file_extensions(),
            'skip_extensions': self.search_criteria_tab.get_skip_extensions(),
            'search_inside_zip': self.search_criteria_tab.search_inside_zip_checkbox.isChecked()  # Add this line
        }

        # Pass the fetched directories to the thread
        self.thread = DuplicatesFinderThread(directories, search_criteria, self.get_hash)
        self.thread.progress_signal.connect(self.update_progress_bar)
        self.thread.finished_signal.connect(self.on_search_complete)
        self.thread.start()

        self.progress_label.show()  # Initially show the "Searching..." label
        self.flashing_timer.start()  # Start flashing when search starts

        self.tree.sortByColumn(0, Qt.AscendingOrder)
    
    def update_progress_bar(self, value):
        self.progress.setValue(value)

    def on_search_complete(self, duplicates):
        # Update the GUI based on the results
        self.tree.clear()

        if not duplicates:
            QMessageBox.information(self, "Info", "No duplicate files found.")
        else:
             for i, files in enumerate(duplicates, start=1):
                for file in files:
                    if "_inside_zip" in file:
                        # Extract the ZIP path and the file inside the ZIP
                        zip_path, inner_file = file.split("_inside_zip/")
                        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                            file_size = "{:.1f}".format(zip_ref.getinfo(inner_file).file_size / 1024)  # Get file size in KB from ZIP metadata
                    else:
                        file_size = "{:.1f}".format(os.path.getsize(file) / 1024)  # Get file size in KB for regular files

                    filename = os.path.basename(file)
                    folder_path = os.path.dirname(file)
                    item = NumericTreeWidgetItem(self.tree)
                    item.setTextAlignment(3, Qt.AlignRight)
                    item.setText(0, filename)
                    #item.setText(1, str(i))
                    item.setData(1, Qt.DisplayRole, i)
                    item.setText(2, folder_path)
                    item.setText(3, file_size)
                    self.update_file_counts()
     
        # Calculate the space that would be freed
        space_to_free = 0
        for files in duplicates:
            for file in files[:-1]:
                if "_inside_zip" in file:
                    zip_path, inner_file = file.split("_inside_zip/")
                    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                        space_to_free += zip_ref.getinfo(inner_file).file_size
                else:
                    space_to_free += os.path.getsize(file)


        # Convert space_t   o_free from bytes to more readable units (like MB or GB)
        space_to_free_MB = space_to_free / (1024 * 1024)  # Convert to MB

        # Update the label with the calculated space
        self.space_to_free_label.setText(f"Space to be freed: {space_to_free_MB:.2f} MB")
     
        self.progress.hide()
        self.find_button.setText("Find Duplicate Files")
        self.flashing_timer.stop()  # Stop flashing when search ends
        self.progress_label.hide()  # Optionally hide the label after search ends
        self.tree.sortByColumn(1, Qt.AscendingOrder)
    def compute_duplicates(self, directories):
        file_dict = {}
        hashes = {}
        print("Hashes computed:", hashes)
        
        for directory in directories:
            for foldername, subfolders, filenames in os.walk(directory):
                for filename in filenames:
                    path = os.path.join(foldername, filename)
                    file_dict[path] = self.get_hash(path)

        for file, hash in file_dict.items():
            if hash not in hashes:
                hashes[hash] = []
            hashes[hash].append(file)
        self.hashes = hashes
        return hashes
        
        print("Self.hashes:", self.hashes)

    def clear_treeview(self):
        self.tree.clear()
        self.update_file_counts()

    def deselect_all(self):
        self.tree.clearSelection()
    def delete_selected(self):

        print("Delete Selected button clicked")
        print("Self.hashes before deletion:", self.hashes)

        selected_items = self.tree.selectedItems()
        selected_count = len(selected_items)

        if selected_count == 0:
            print("No items selected for deletion.")
            return

        selected_option_id = self.delete_options_tab.button_group.checkedId()
        print(f"Selected option ID: {selected_option_id}")

        # Loop through selected items and delete files  
        for item in selected_items:
            
            file_path = os.path.join(item.text(2), item.text(0))

            if "_inside_zip" in file_path:
                self.delete_file_from_zip(file_path) 
            else:
                if selected_option_id == 1:
                    self.delete_file_permanently(file_path)
                elif selected_option_id == 2:
                    self.move_file_to_trash(file_path)
                elif selected_option_id == 3:
                    self.move_file_to_new_folder(file_path, item)
                elif selected_option_id == 4:
                    self.replace_file_with_hardlink(file_path, item)

        self.check_remaining_files()
        self.update_file_counts()
        self.tree.sortByColumn(1, Qt.AscendingOrder)


    def delete_file_from_zip(self, file_path):
        zip_path, inner_file = file_path.split("_inside_zip/")
        try:
            subprocess.run([r'C:\Program Files\7-Zip\7z.exe', 'd', zip_path, inner_file], check=True)
        except (PermissionError, subprocess.CalledProcessError, FileNotFoundError, Exception) as e:
            self.handle_deletion_error(e, file_path)
        else:
            # Adding logging here
            #deleted_id = item.text(1)  # get id from deleted item
            print(f"Deleted file from ZIP with ID: {deleted_id}")
            if deleted_id in self.hashes:
                print(f"Paths associated with this ID before removal: {self.hashes[deleted_id]}")
            else:
                print(f"Hash {deleted_id} not found in records.")
            self.remove_item_from_tree(item)


    def delete_file_permanently(self, file_path):
        try:
            os.remove(file_path)
        except OSError as e:
            self.handle_delete_error(e, file_path)
      
        # Update hash lookup
        hash = get_hash(file_path) 
        self.files_by_hash[hash].remove(file_path)

        if len(self.files_by_hash[hash]) == 1:
            # No longer duplicate, remove hash
            del self.files_by_hash[hash]

    def move_file_to_trash(self, file_path, item):
        try:
            send2trash.send2trash(file_path)
        except Exception as e:
            self.handle_deletion_error(e, file_path)
        else:
            self.remove_item_from_tree(item)

    def move_file_to_new_folder(self, file_path, item):
        new_folder = self.delete_options_tab.new_folder_entry.text()
        if not new_folder:
            QMessageBox.critical(self, "Error", "No folder specified")
            return
        os.makedirs(new_folder, exist_ok=True)
        try:
            os.rename(file_path, os.path.join(new_folder, item.text(0)))
        except OSError as e:
            self.handle_deletion_error(e, file_path)
        else:
            self.remove_item_from_tree(item)

    def replace_file_with_hardlink(self, file_path, item):
        selected_paths = [os.path.join(item.text(2), item.text(0)) for item in self.tree.selectedItems()]
        
        for group in self.hashes.values():
            if len(group) < 2:
                continue
            
            primary_file = group[0]
            if primary_file in selected_paths:
                continue
            
            for duplicate_file in group[1:]:
                if duplicate_file not in selected_paths:
                    continue
                    
                try:
                    os.remove(duplicate_file)  # Remove the duplicate
                    os.link(primary_file, duplicate_file)  # Create a hardlink from primary to duplicate's location
                except OSError as e:
                    QMessageBox.critical(self, "Error", f"Could not replace {duplicate_file} with a hardlink to {primary_file}: {str(e)}")
                else:
                    self.remove_item_from_tree(item)
                    print(f"Replaced {duplicate_file} with a hardlink to {primary_file}")

    def handle_deletion_error(self, error, file_path):
        QMessageBox.critical(self, "Error", f"Could not process {file_path}: {str(error)}")

    def remove_item_from_tree(self, item):
        index = self.tree.indexOfTopLevelItem(item)
        self.tree.takeTopLevelItem(index)


    
    def check_remaining_files(self):
        hashes = {}

        # Recompute hashes for remaining files
        for i in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(i)
            file_path = os.path.join(item.text(2), item.text(0))  # Get full path of file
            if os.path.exists(file_path):  # If file still exists
                file_hash = self.get_hash(file_path)
                if file_hash not in hashes:
                    hashes[file_hash] = []
                hashes[file_hash].append(item)

        # Check for files that are no longer duplicates
        for hash, items in hashes.items():
            if len(items) <= 1:  # If file is no longer a duplicate
                for item in items:
                    index = self.tree.indexOfTopLevelItem(item)
                    self.tree.takeTopLevelItem(index)  # Remove item from tree
                    
    def update_file_counts(self):
        # Count the total files and duplicate groups
        total_files = self.tree.topLevelItemCount()
        duplicate_groups = len(set([self.tree.topLevelItem(i).text(1) for i in range(total_files)]))

        # Update the labels
        self.total_files_label.setText(f"Total files: {total_files}")
        self.duplicate_files_label.setText(f"Duplicate files: {duplicate_groups}")
        
    def move_selected_to_directory(self):
        directory = QFileDialog.getExistingDirectory(self, "Select Directory")
        if not directory:
            return
        selected_items = self.tree.selectedItems()
        for item in selected_items:
            file_path = os.path.join(item.text(2), item.text(0))  # Get full path of file
            try:
                os.rename(file_path, os.path.join(directory, item.text(0)))  # Move file to directory
            except OSError as e:
                QMessageBox.critical(self, "Error", f"Could not move file {file_path}: {str(e)}")
            else:
                index = self.tree.indexOfTopLevelItem(item)
                self.tree.takeTopLevelItem(index)  # Remove item from tree

    def move_selected_to_new_folder(self):
        directory = QFileDialog.getExistingDirectory(self, "Select Directory")
        if not directory:
            return
        new_folder, ok = QInputDialog.getText(self, "New Folder", "Enter new folder name:")
        if not ok or not new_folder:
            return
        os.makedirs(os.path.join(directory, new_folder), exist_ok=True)
        self.move_selected_to_directory()   
    
    
    def get_hash(self, file_path, block_size=65536):
        # Create a hash for the file
        file_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            for block in iter(lambda: f.read(block_size), b""):
                file_hash.update(block)
        return file_hash.hexdigest()

    def update_duplicates(self, deleted_item):
        print("Self.hashes on update:", self.hashes)
        deleted_id = deleted_item.text(1) # get id from deleted item
        print("Deleted ID:", deleted_id)
        
        if deleted_id in self.hashes and len(self.hashes[deleted_id]) == 1:
#        if len(self.hashes[deleted_id]) == 1:
        # Remove hash and rows with this id
            self.hashes.pop(deleted_id)  
            self.remove_id_rows(deleted_id)
        else:
        # Recompute as before
            self.tree.clear()
            hashes = self.compute_duplicates(self.directories)
            self.populate_tree(hashes)

    def remove_id_rows(self, id):

      # Find all matching items
      matches = self.tree.findItems(id, Qt.MatchExactly, 1)
      
      # Remove the matching items
      for item in matches:
        self.tree.takeTopLevelItem(self.tree.indexOfTopLevelItem(item))
        
    def save_selection(self, selected, deselected):
        print("Rows selected:", [index.row() for index in self.tree.selectionModel().selectedRows()])

    def clear_selection(self, selected, deselected):
        print("Rows deselected:", [index.row() for index in deselected.indexes()])

    def export_to_excel(self):
        # Create a new workbook and select active sheet
        wb = Workbook()
        ws = wb.active

        # Write headers
        for col_num, header in enumerate(self.tree.setHeaderLabels(), 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = Font(bold=True)

        # Write data to the sheet
        for row_num, tree_item in enumerate(self.iterate_tree_items(self.tree), 2):
            for col_num, data in enumerate(tree_item, 1):
                ws.cell(row=row_num, column=col_num, value=data)

        # Ask user where to save the Excel file
        options = QFileDialog.Options()
        filePath, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if filePath:
            if not filePath.endswith('.xlsx'):
                filePath += '.xlsx'
            wb.save(filePath)

    def iterate_tree_items(self, tree):
        """Yield each item in the QTreeWidget as a list of its data."""
        root = tree.invisibleRootItem()
        for i in range(root.childCount()):
            item = root.child(i)
            yield [item.text(col) for col in range(item.columnCount())]

    def toggle_searching_visibility(self):
        """Toggle the visibility of the 'Searching...' label."""
        if self.progress_label.isVisible():
            self.progress_label.hide()
        else:
            self.progress_label.show()



#############################
# right click menu
#############################

    def handle_context_menu(self, position):
        # Create context menu
        menu = QMenu(self.tree)
        select_same_folder_action = menu.addAction("Select files in same folder")
        select_same_folder_action.triggered.connect(self.select_files_in_same_folder)
        select_similar_folder_action = menu.addAction("Select files in similar folder")
        select_similar_folder_action.triggered.connect(self.select_files_in_similar_folder)
        select_same_drive_action = menu.addAction("Select files on same drive")
        select_same_drive_action.triggered.connect(self.select_files_on_same_drive)
        select_duplicated_elsewhere_action = menu.addAction("Select files that duplicate files in this file's folder elsewhere")
        select_duplicated_elsewhere_action.triggered.connect(self.select_files_duplicated_elsewhere)
            
        open_file_action = menu.addAction("Open file")
        open_file_action.triggered.connect(self.open_file)

        open_file_location_action = menu.addAction("Open file location")
        open_file_location_action.triggered.connect(self.open_file_location)

        menu.exec_(self.tree.viewport().mapToGlobal(position))

    def select_files_in_same_folder(self):
        selected_items = self.tree.selectedItems()
        if not selected_items:
            return
        selected_folder = selected_items[0].text(2)
        self.tree.clearSelection()
        for i in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(i)
            if item.text(2) == selected_folder:
                item.setSelected(True)

    def select_files_in_similar_folder(self):
        selected_items = self.tree.selectedItems()
        if not selected_items:
            return
        selected_folder = selected_items[0].text(2)
        
        dialog = QInputDialog(self)
        dialog.setInputMode(QInputDialog.TextInput)
        dialog.setWindowTitle("Select Similar Folder")
        dialog.setLabelText("Enter similar folder:")
        dialog.setTextValue(selected_folder)
        dialog.resize(500, 100)  # Set the desired width and height
        
        ok = dialog.exec_()
        similar_folder = dialog.textValue()
        
        if ok and similar_folder:
            self.tree.clearSelection()
            for i in range(self.tree.topLevelItemCount()):
                item = self.tree.topLevelItem(i)
                if similar_folder in item.text(2):
                    item.setSelected(True)

    def open_file(self):
        selected_items = self.tree.selectedItems()
        if not selected_items:
            return
        file_path = os.path.join(selected_items[0].text(2), selected_items[0].text(0))  # Get full path of file
        if sys.platform == "win32":
            os.startfile(file_path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", file_path])
        else:
            subprocess.Popen(["xdg-open", file_path])

    def open_file_location(self):
        selected_items = self.tree.selectedItems()
        if not selected_items:
            return
        file_location = selected_items[0].text(2)  # Get file location
        if sys.platform == "win32":
            os.startfile(file_location)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", file_location])
        else:
            subprocess.Popen(["xdg-open", file_location])  

    def open_file(self):
        selected_items = self.tree.selectedItems()
        if not selected_items:
            return
        file_path = os.path.join(selected_items[0].text(2), selected_items[0].text(0))  # Get full path of file
        if sys.platform == "win32":
            os.startfile(file_path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", file_path])
        else:
            subprocess.Popen(["xdg-open", file_path])

    def select_one_file_per_group(self):
        duplicate_ids = set()
        self.tree.clearSelection()
        for i in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(i)
            if item.text(1) not in duplicate_ids:
                duplicate_ids.add(item.text(1))
            else:
                item.setSelected(True)
    
    def select_files_on_same_drive(self):
        selected_items = self.tree.selectedItems()
        if not selected_items:
            return
        selected_drive = os.path.splitdrive(selected_items[0].text(2))[0]  # Get the drive of the selected file
        self.tree.clearSelection()
        for i in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(i)
            if os.path.splitdrive(item.text(2))[0] == selected_drive:  # If the drive of the file matches the selected drive
                item.setSelected(True)

    def select_files_duplicated_elsewhere(self):
        selected_items = self.tree.selectedItems()
        if not selected_items:
            return
        selected_folder = selected_items[0].text(2)  # Get the folder of the selected file
        self.tree.clearSelection()
        for i in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(i)
            if item.text(2) != selected_folder:  # If the folder of the file does not match the selected folder
                item.setSelected(True)
  
  
class SearchDirectoriesTab(QWidget):
    def __init__(self, *args, **kwargs):
        super(SearchDirectoriesTab, self).__init__(*args, **kwargs)

        main_layout = QVBoxLayout()

        # Create the go up button
        self.go_up_button = QPushButton("Go to parent folder")
        self.go_up_button.clicked.connect(self.go_up)
        self.go_up_button.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.Fixed)
        self.go_up_button.setMaximumWidth(200)  # Set the maximum width to 200 pixels

        # Add the Go Up button to the main layout
        main_layout.addWidget(self.go_up_button)

        # Add a Directory Tree View
        self.file_system_model = QFileSystemModel()
        self.file_system_model.setRootPath(QDir.rootPath())
        self.file_system_model.setFilter(QDir.NoDotAndDotDot | QDir.AllDirs)  # To display only directories

        self.tree = QTreeView()
        self.tree.setModel(self.file_system_model)
        self.tree.setRootIndex(self.file_system_model.index(QDir().homePath() + "/Documents"))
        self.tree.doubleClicked.connect(self.add_directory_from_tree)

        # Hide the unnecessary columns
        self.tree.hideColumn(1)  # Hide 'Size' column
        self.tree.hideColumn(2)  # Hide 'Type' column
        self.tree.hideColumn(3)  # Hide 'Date Modified' column

        # Create the table for displaying directories
        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["Scan Against Self", "Folder"])
        self.table.horizontalHeader().setStretchLastSection(True)

        # Create a splitter to manage the layout of tree and table
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(self.tree)
        splitter.addWidget(self.table)
        splitter.setStretchFactor(0, 1)  # Set the stretch factor for tree
        splitter.setStretchFactor(1, 2)  # Set the stretch factor for table

        # Add the splitter to the main layout
        main_layout.addWidget(splitter)

        # Create the buttons
        self.add_directory_button = QPushButton("Add Directory")
        self.add_directory_button.clicked.connect(self.add_directory)
        self.remove_directory_button = QPushButton("Remove Directory")
        self.remove_directory_button.clicked.connect(self.remove_directory)

        # Set size policy for these buttons to be Fixed
        self.add_directory_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.remove_directory_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        # Create a horizontal layout for the buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch(1)  # Add a stretchable space before the buttons
        button_layout.addWidget(self.add_directory_button)
        button_layout.addWidget(self.remove_directory_button)

        # Add the buttons layout to the main layout
        main_layout.addLayout(button_layout)

        self.setLayout(main_layout)

    def go_up(self):
        index = self.tree.rootIndex()
        parent_index = index.parent()
        if parent_index.isValid():  # Make sure the parent directory exists
            self.tree.setRootIndex(parent_index)

    def add_directory_from_tree(self, index):
        path = self.file_system_model.filePath(index)
        print(f"Selected path: {path}")

        # Add the selected directory to the table
        row = self.table.rowCount()
        self.table.insertRow(row)
        checkbox = QCheckBox()
        self.table.setCellWidget(row, 0, checkbox)
        self.table.setItem(row, 1, QTableWidgetItem(path))

    def add_directory(self):
        directory = QFileDialog.getExistingDirectory(self, "Select Directory")
        if directory:  # If a directory is selected
            row = self.table.rowCount()
            self.table.insertRow(row)
            checkbox = QCheckBox()
            self.table.setCellWidget(row, 0, checkbox)
            self.table.setItem(row, 1, QTableWidgetItem(directory))

    def remove_directory(self):
        for item in self.table.selectedItems():
            self.table.removeRow(item.row())

    def get_directories(self):
        # Get all directories from the QTableWidget
        directories = []
        for row in range(self.table.rowCount()):
            directory = self.table.item(row, 1).text()
            scan_self = self.table.cellWidget(row, 0).isChecked()
            directories.append((directory, scan_self))
        return directories

class SearchCriteriaTab(QWidget):
    def __init__(self, *args, **kwargs):
        super(SearchCriteriaTab, self).__init__(*args, **kwargs)

        layout = QVBoxLayout()

        # Add a spacer to push the contents to the top
        layout.addItem(QSpacerItem(20, 50, QSizePolicy.Minimum))

        # Minimum file size field
        min_file_size_layout = QHBoxLayout()
        min_file_size_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Fixed, QSizePolicy.Fixed))  # Add a spacer item
        min_file_size_label = QLabel("Minimum file size (KB):")
        min_file_size_layout.addWidget(min_file_size_label)
        self.min_file_size_entry = QLineEdit()
        min_file_size_layout.addWidget(self.min_file_size_entry)
        layout.addLayout(min_file_size_layout)

        # Maxmimumfile size field
        max_file_size_layout = QHBoxLayout()
        max_file_size_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Fixed, QSizePolicy.Fixed))  # Add a spacer item
        max_file_size_label = QLabel("Maxmimum file size (KB):")
        max_file_size_layout.addWidget(max_file_size_label)
        self.max_file_size_entry = QLineEdit()
        max_file_size_layout.addWidget(self.max_file_size_entry)
        layout.addLayout(max_file_size_layout)


        # File extensions field
        file_extensions_layout = QHBoxLayout()
        file_extensions_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Fixed, QSizePolicy.Fixed))  # Add a spacer item
        file_extensions_label = QLabel("File extensions (comma-separated):")
        file_extensions_layout.addWidget(file_extensions_label)
        self.file_extensions_entry = QLineEdit()
        file_extensions_layout.addWidget(self.file_extensions_entry)
        layout.addLayout(file_extensions_layout)

        # Extensions to skip field
        skip_extensions_layout = QHBoxLayout()
        skip_extensions_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Fixed, QSizePolicy.Fixed))
        skip_extensions_label = QLabel("Extensions to skip (comma-separated):")
        skip_extensions_layout.addWidget(skip_extensions_label)
        self.skip_extensions_entry = QLineEdit()
        skip_extensions_layout.addWidget(self.skip_extensions_entry)
        layout.addLayout(skip_extensions_layout)

        layout.addSpacing(20)
        
        # Checkbox for searching inside ZIP files
        self.search_inside_zip_checkbox = QCheckBox("Search inside ZIP files [requires 7-zip to delete files within the zip file]")
        self.search_inside_zip_checkbox.setChecked(True)  # Set the default state to checked

        # QHBoxLayout for indenting the checkbox
        zip_search_layout = QHBoxLayout()
        zip_search_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Fixed, QSizePolicy.Fixed))  # Small indentation
        zip_search_layout.addWidget(self.search_inside_zip_checkbox)  # Add the checkbox
        zip_search_layout.addStretch()  # Add stretch to take any remaining space

        layout.addLayout(zip_search_layout)  # Add the QHBoxLayout to the main layout



        # Space between radio buttons
        layout.addSpacing(50)

        percent_similar_layout = QHBoxLayout()
        percent_similar_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Fixed, QSizePolicy.Fixed))  # Add a spacer item
        percent_similar_label = QLabel("Percent Similar (in progress, not currently working):")
        percent_similar_layout.addWidget(percent_similar_label)
        self.percent_similar_entry = QLineEdit()
        self.percent_similar_entry.setEnabled(False)  # Disable the input field until the feature is implemented
        percent_similar_layout.addWidget(self.percent_similar_entry)
        layout.addLayout(percent_similar_layout)

        # Add a spacer to push the contents to the top
        layout.addItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding))

        self.setLayout(layout)


    def get_min_file_size(self):
        # Return the entered minimum file size in KB, or 0 if no size is entered
        return int(self.min_file_size_entry.text()) if self.min_file_size_entry.text().isdigit() else 0

    def get_max_file_size(self):
        # Return the entered maximum file size in KB, or sys.maxsize if no size is entered
        return int(self.max_file_size_entry.text()) if self.max_file_size_entry.text().isdigit() else sys.maxsize

    def get_file_extensions(self):
        # Return the entered file extensions as a list, or an empty list if no extensions are entered
        return [ext.strip() for ext in self.file_extensions_entry.text().split(",")] if self.file_extensions_entry.text() else []

    def get_skip_extensions(self):
        # Return the entered skip extensions as a list, or an empty list if no extensions are entered
        return [ext.strip() for ext in self.skip_extensions_entry.text().split(",")] if self.skip_extensions_entry.text() else []
    
    def get_percent_similar(self):
        # Placeholder method to be implemented in the future
        pass
        
    def should_search_inside_zip(self):
        """Return True if the user wants to search inside ZIP files, otherwise return False."""
        return self.search_inside_zip_checkbox.isChecked()        
        
class DeleteOptionsTab(QWidget):
    def __init__(self, *args, **kwargs):
        super(DeleteOptionsTab, self).__init__(*args, **kwargs)

        layout = QVBoxLayout()
        layout.setSpacing(5)

        # Add a spacer to push the contents to the top
        layout.addItem(QSpacerItem(20, 50, QSizePolicy.Minimum))

        # Radio buttons for choosing the delete method
        self.button_group = QButtonGroup()

        self.delete_radio = QRadioButton("Delete duplicates permanently - DOES NOT SEND FILES TO TRASH")
        self.button_group.addButton(self.delete_radio, 1)

        delete_radio_layout = QHBoxLayout()
        delete_radio_layout.addSpacing(40)  # Indent by 20 pixels
        delete_radio_layout.addWidget(self.delete_radio)
        layout.addLayout(delete_radio_layout)

        # Space between radio buttons
        layout.addSpacing(10)

        self.move_radio = QRadioButton("Move duplicates to Trash")
        self.button_group.addButton(self.move_radio, 2)

        move_radio_layout = QHBoxLayout()
        move_radio_layout.addSpacing(40)  # Indent by 20 pixels
        move_radio_layout.addWidget(self.move_radio)
        layout.addLayout(move_radio_layout)

        # Space between radio buttons
        layout.addSpacing(10)

        # Add hardlinks radio button
        self.hardlink_radio = QRadioButton("Add hardlinks for duplicates")
        self.button_group.addButton(self.hardlink_radio, 4)

        hardlink_radio_layout = QHBoxLayout()
        hardlink_radio_layout.addSpacing(40)  # Indent by 20 pixels
        hardlink_radio_layout.addWidget(self.hardlink_radio)
        layout.addLayout(hardlink_radio_layout)

        # Space between radio buttons
        layout.addSpacing(10)

        # Combine the radio button and input field in a single horizontal layout
        combined_layout = QHBoxLayout()
        combined_layout.addSpacing(40)  # Indent by 20 pixels

        self.new_folder_radio = QRadioButton("Move duplicates to a new folder:")
        self.button_group.addButton(self.new_folder_radio, 3)
        combined_layout.addWidget(self.new_folder_radio)

        self.new_folder_entry = QLineEdit()
        combined_layout.addWidget(self.new_folder_entry)

        self.select_folder_button = QPushButton("Select Folder")
        self.select_folder_button.clicked.connect(self.select_folder)
        combined_layout.addWidget(self.select_folder_button)

        layout.addLayout(combined_layout)

        # Add a spacer to push the contents to the top
        layout.addItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding))

        # Set "Delete duplicates permanently" as the default option
        self.delete_radio.setChecked(True)

        self.setLayout(layout)
        # Create a label for the folder input field
        self.new_folder_label = QLabel("Folder for moving/hardlinks:")
        layout.addWidget(self.new_folder_label)

    def select_folder(self):
        directory = QFileDialog.getExistingDirectory(self, "Select Directory")
        if directory:  # If a directory is selected
            self.new_folder_entry.setText(directory)  # Set it in the input field
class HelpTab(QWidget):
    def __init__(self, version, *args, **kwargs):
        super(HelpTab, self).__init__(*args, **kwargs)

        layout = QVBoxLayout()

        # Create a QTextEdit widget to display multi-line text
        self.text_edit = QTextEdit()
        self.text_edit.setReadOnly(True)  # Make the text read-only

        # Set the font size
        font = QFont()
        font.setPointSize(14)
        self.text_edit.setFont(font)

        # Add some text to the QTextEdit widget
        # Note: We are using HTML formatting for the email link
        self.text_edit.setHtml("""

This is version {version} and it still a beta level software...<br><br>

The following features are not yet active:
<div style="font-size:10pt;">
&nbsp;&nbsp;&nbsp;&nbsp;a) limit scanning of folder against itself<br>
&nbsp;&nbsp;&nbsp;&nbsp;b) percent similar function - this will allow for scanning for similar files that are not exact matches<br>
&nbsp;&nbsp;&nbsp;&nbsp;c) limit search by date<br>
&nbsp;&nbsp;&nbsp;&nbsp;e) limit search by same/similar file names<br>
&nbsp;&nbsp;&nbsp;&nbsp;f) limit search to same created/modified date<br>
&nbsp;&nbsp;&nbsp;&nbsp;g) many, many, many more...<br><br><br>

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
""")


        layout.addWidget(self.text_edit)

        self.setLayout(layout)





if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
